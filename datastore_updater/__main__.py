# -------
# imports

import os
import os.path as osp
import sys

import numpy as np
import openpyxl as op
import pandas as pd
import selenium.common.exceptions as sncmer
import selenium.webdriver as snwd

from .API import sg_yahoo_finance as api
from .._notifier import main as _nf


# ---------
# supp func

def _driverinit(webdriver_boo, webdriver_strlst, main_dir):

    # initialising webdriver
    driver = None
    if webdriver_boo:
        print('\ninitialising webdriver...')
        init_boo = False
        for webdriver_str in webdriver_strlst:
            driver_dir = main_dir + '/_webdriver/' + webdriver_str
            chrome_options = snwd.ChromeOptions()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--disable-gpu')

            ## for bug in selenium/rasp
            if webdriver_str == 'chromedriver_rasp32_74':
                chrome_options.add_argument(
                    '--disable-features=VizDisplayCompositor')

            try:                # driver OS is correct
                driver = snwd.Chrome(driver_dir,
                                     options=chrome_options)
                init_boo = True
            except (OSError, sncmer.SessionNotCreatedException) as e:
                # wrong OS or wrong version, trying the next one in lst
                continue

            if init_boo:
                print('webdriver is initialised as {}'.
                      format(webdriver_str))
                break

        if not init_boo:        # fail to init, outdated webdriver
            msg = \
                '''<pre>project: invest_scripts
                event  : Depreciation
                error  : webdriver outdated
                </pre>'''
            _nf('investtianbot', 'markdown', msg, sendclient_boo=False)

            raise DeprecationWarning('chromedriver outdated')

    return driver


# ---------
# operation
'''
purpose is to update the 'Stocks' sheet in 'datastore.xlsx'
'''

def main(update_file=None, verb_boo=True):
    '''
    update_file:: str, dir, if given, update will be performed on given
                                      file
                          , else, update will be performed on default
                                  store_file
    '''

    # ------
    # params
    sheet = 'Stock'
    excolname_str = 'Stock EX'
    codecolname_str = 'Stock Code'
    namecolname_str = 'Stock Name'
    keywordstart_ind = 3        # skipping EX, code and name
    main_dir = osp.dirname(osp.dirname(osp.abspath(__file__)))

    webdriver_boo = True
    webdriver_strlst = list(filter(
        lambda x: 'chromedriver' in x, os.listdir(main_dir + '/_webdriver')
    ))

    if not update_file:
        store_file = osp.dirname(main_dir) + '/' + 'datastore.xlsx'
    else:
        store_file = update_file


    # ---------
    # operation

    if not verb_boo:            # disabling print
        sys.stdout = open(os.devnull, 'w')
    print('\n--- beginning {} update... ---'.format(store_file))

    # reading file
    print('\nloading workbook, working on worksheet: {}...'.\
          format(sheet))
    df = pd.read_excel(store_file, sheet).fillna('')
    ## sorting file
    lst = list(df.to_numpy())
    lst.sort(key=lambda x: x[keywordstart_ind-1].lower())
    df = pd.DataFrame(np.array(lst), columns=df.columns.values)


    # performing pull
    print('\npulling data...')
    ex_ara = df[excolname_str].values
    code_ara = df[codecolname_str].values
    name_ara = df[namecolname_str].values
    keyword_ara = df.columns.values[keywordstart_ind:]
    update_df = df.copy()
    for i in range(len(code_ara)):
        print('retrieving for: {} {}'.format(code_ara[i], name_ara[i]))

        # initialising data in each loop
        driver = _driverinit(webdriver_boo, webdriver_strlst, main_dir)

        # perform pull
        if webdriver_boo:
            pulldata_lst = api.func(ex_ara[i], code_ara[i], keyword_ara,
                                    driver)
        else:
            pulldata_lst = api.func(ex_ara[i], code_ara[i], keyword_ara)

        # assimilate data
        update_df.loc[i:i, keyword_ara[0]:keyword_ara[-1]] =\
            pulldata_lst

    # pasting data
    ## clearing old sheets; by removing and adding
    ## This is done to remove any old stocks that were not removed bcos
    ## no. of current stocks is fewer
    wb = op.load_workbook(store_file)
    for sheet in [sheet]:
        wb.remove(wb[sheet])
        wb.create_sheet(sheet)
    wb.save(store_file)
    ## writing file
    excelBook = op.load_workbook(store_file)
    with pd.ExcelWriter(store_file, engine='openpyxl') as writer:
        writer.book = excelBook
        writer.sheets = dict((ws.title, ws)\
                             for ws in excelBook.worksheets)
        update_df.to_excel(writer, sheet
                           , header=True, index=False)
        writer.save()


    print('\n--- {} update complete! ---'.format(store_file))
    if not verb_boo:                # enabling print again
        sys.stdout = sys.__stdout__


# ----------
# executable

if __name__ == '__main__':
    main()

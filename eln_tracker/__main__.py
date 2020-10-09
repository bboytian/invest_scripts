'''
This script is intended to be run on the raspberry pi server
as it has to check with previous records

ELN logic:

1. Knock In (KI)

 - if any stock < 85% of stock original price at any given point in time

 - given KI, if any stock < 95% of original price by maturity date, must but worst performing stock for strike price (95%)

 - once knock in for any stock in the ELN, notify: which ELN, stock KI and price

 - do not notify KI anymore

 - notify when matures

2. Knock Out (KO)

 - if all 4 stocks >= 100% of original stock price, ELN ends

 - notify which ELN, and stock prices

 - do not track ELN anymore; can mark as matured?

3. Maturity date

 - notify which ELN, and stock prices, and KI boolean, if KI, notify which stock has to be purchased
'''
# -------
# imports

import datetime as dt
import os.path as osp

import numpy as np
import openpyxl as op
import pandas as pd

from . import sendmsg_func
from .. import datastore_updater


# ---------
# operation
'''
tracking of KI/KO/maturity events for ELN:

0. needs logic for KI/KO

1. needs to keep a record of previous KI/KO events

2. needs to notify via telegram bot
'''
if __name__ == '__main__':

    # ------
    # params

    # document names
    eln_sheet = 'ELN'
    stock_sheet = 'Stock'
    pkg_dir =  osp.dirname(osp.abspath(__file__))
    store_file = osp.dirname(osp.dirname(pkg_dir)) + '/' \
        + 'datastore.xlsx'
    rec_file = osp.dirname(pkg_dir) + '/_records/' \
        + 'elnrecord.xlsx'
    
    # ELN sheet formatting; for elnrecord.xlsx and datastore.xlsx
    coloff_int = 2
    boorowoff_int = 3
    stockrowoff_int = 9
    codecol = 'ELN XS'
    elnid_row = 'ELN XS'
    account_row = 'Account'
    activedate_row = 'First Observation Date'
    maturedate_row = 'Maturity Date'
    kiboo_row = 'KI boolean'
    matureboo_row = 'Mature boolean'     # ko and mature are equiv
    kofactor_row = 'Knock Out Factor'    
    kifactor_row = 'Knock In Factor'
    strikefactor_row = 'Strike Factor'

    # ELN records update
    elnreccol_lst = [kiboo_row, matureboo_row]
    
    # Stock sheet formatting
    scodecol = 'Stock Code'
    spricecol = 'Price'
    
    
    # ---------
    # operation

    print('\n--- begin ELN tracking for {}... ---'.format(store_file))

    
    print('\nupdating ELN records with datastore... ')
    # reading old files
    ## dat => datastore, rec => elcrecords
    dat_df = pd.read_excel(store_file, eln_sheet, index_col=codecol)
    rec_df = pd.read_excel(rec_file, eln_sheet, index_col=codecol)

    
    # updating ELN
    ## getting new ELN
    newrec_df = dat_df.copy()   # to be written over ELN records
    datcol_ind = dat_df.columns[coloff_int:]
    for datcol in datcol_ind:
        reccol = 'rec: {}'.format(datcol)
        newrec_df.insert(datcol_ind.get_loc(datcol)*2+coloff_int,
                         reccol, np.nan)

    ## filling in previous ELN data
    newreccol_ind = newrec_df.columns[coloff_int:]
    for newrecrow in elnreccol_lst:
        for newreccol in newreccol_ind:
            try:                # accounts for old ELN data
                newrec_df.loc[newrecrow, newreccol] = \
                    rec_df.loc[newrecrow, newreccol]
            except KeyError:    # accounts for removed ELN
                pass

            
    # writing over records
    rec_df = newrec_df

        
    print('\ncomparing ELN records with stock prices...')
    sdat_df = pd.read_excel(store_file, stock_sheet, # 's' for stock
                            index_col=scodecol)

    
    for elnid in datcol_ind:
        print('\nworkign on ELN: {}'.format(elnid))

    # retrieving record information        
        reccol_sr = rec_df[elnid] 

        elnid_str = elnid
        account_str = reccol_sr[account_row]

        kofactor_flt = reccol_sr[kofactor_row]
        kifactor_flt = reccol_sr[kifactor_row]
        strikefactor_flt = reccol_sr[strikefactor_row]

        sreccol_sr = reccol_sr[stockrowoff_int:] # 's' for stock
        sreccol_sr = sreccol_sr[~sreccol_sr.isna()] # filtered 'NA'
        sdatcol_sr = sdat_df[spricecol][sreccol_sr.index]
        
    # retrieving new booleans

        ## booleans for time
        active_ts = pd.Timestamp(reccol_sr[activedate_row])
        mature_ts = pd.Timestamp(reccol_sr[maturedate_row])
        now_ts = pd.Timestamp(dt.datetime.now())
        active_boo = active_ts <= now_ts
        mature_boo = mature_ts <= now_ts

        ## booleans for KI
        kiboo_sr = (sdatcol_sr < kifactor_flt*sreccol_sr)
        ki_boo = kiboo_sr.any()
        prevki_boo = reccol_sr[kiboo_row]
        if np.isnan(prevki_boo): # first time checking ELN
            rec_df.loc[kiboo_row, elnid] = 0 # False for pandas to read
            prevki_boo = 0


        ## boolean for KO
        koboo_sr = (sdatcol_sr >= kofactor_flt*sreccol_sr)
        ko_boo = koboo_sr.all()

        ## boolean for mature
        prevmature_boo = reccol_sr[matureboo_row]
        if np.isnan(prevmature_boo): # first time checking ELN
            rec_df.loc[matureboo_row, elnid] = 0
            prevmature_boo = 0

        ## boolean for message sending
        sendmsg_boo = False
        event_str = ''

        
    # handling logic
    ## '+' => may have to buy worst performing stock
    ## which depends on stirke factor and calculated in sendmsg_func
        if active_boo and\
             not ki_boo and not prevki_boo and\
             mature_boo and not prevmature_boo:
            rec_df.loc[matureboo_row, elnid] = 1
            sendmsg_boo, event_str = True, 'Mature'

        elif active_boo and\
             ki_boo and not prevki_boo and\
             not mature_boo and not prevmature_boo:
            rec_df.loc[kiboo_row, elnid] = 1
            sendmsg_boo, event_str = True, 'KI'

        elif active_boo and\
             ki_boo and not prevki_boo and\
             mature_boo and not prevmature_boo:
            rec_df.loc[kiboo_row, elnid] = 1
            rec_df.loc[matureboo_row, elnid] = 1
            sendmsg_boo, event_str = True, 'Simul Mature & KI+'

        elif (active_boo and\
              ki_boo and prevki_boo and\
              mature_boo and not prevmature_boo)\
              or\
              (active_boo and\
               not ki_boo and prevki_boo and\
               mature_boo and not prevmature_boo):
            rec_df.loc[matureboo_row, elnid] = 1
            sendmsg_boo, event_str = True, 'Mature & prev KI+'

        elif (active_boo and\
              ko_boo and prevki_boo and\
              not mature_boo and not prevmature_boo)\
              or\
              (active_boo and\
              ko_boo and prevki_boo and\
              mature_boo and not prevmature_boo)\
              or\
              (active_boo and\
              ko_boo and not prevki_boo and\
              mature_boo and not prevmature_boo)\
              or\
              (active_boo and\
              ko_boo and not prevki_boo and\
              not mature_boo and not prevmature_boo):
            rec_df.loc[matureboo_row, elnid] = 1 #no harm apply all
            sendmsg_boo, event_str = True, 'KO'


    # Sending msg if needed
        if sendmsg_boo:
            print('sending message...')
            '''finish up function in the other script'''
            sendmsg_func.func(
                event_str,
                elnid_str, account_str, mature_ts, 
                sreccol_sr, sdatcol_sr,
                kofactor_flt, kifactor_flt, strikefactor_flt
            )

    print('\nsaving records in {}'.format(rec_file))
    # clearing old sheets; by removing and adding
    ## This is done to remove any old ELNs that were not removed bcos
    ## no. of current ELNs is fewer
    wb = op.load_workbook(rec_file)
    for sheet in wb.sheetnames:
        try:
            wb.remove(wb[sheet])
        except KeyError:  # if stock_sheet isnt present in old records
            pass
        wb.create_sheet(sheet)
    wb.save(rec_file)
    # writing file
    with pd.ExcelWriter(rec_file, engine='openpyxl') as writer:
        writer.book = wb
        writer.sheets = dict((ws.title, ws)\
                             for ws in wb.worksheets\
                             if ws.title == eln_sheet)
        rec_df.to_excel(writer, eln_sheet,
                        header=True, index=True)
        writer.save()                    

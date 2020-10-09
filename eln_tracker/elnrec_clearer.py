'''
This script acts to empty the elnrecords for a fresh run of eln_updater
'''
# -------
# imports

import os.path as osp

import openpyxl as op


# ------
# params

pkg_dir =  osp.dirname(osp.abspath(__file__))
rec_file = osp.dirname(pkg_dir) + '/_records/' + 'elnrecord.xlsx'


# ---------
# operation

wb = op.load_workbook(rec_file)
for sheet in wb.sheetnames:
    try:
        wb.remove(wb[sheet])
    except KeyError:  # if stock_sheet isnt present in old records
        pass
    wb.create_sheet(sheet)
wb.save(rec_file)
